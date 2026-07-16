import openpyxl, os, math

XLSX      = r'C:\Users\predator\Desktop\new ims\Ims-desktop-application\Product_Master_Merged.xlsx'
OUT_DIR   = r'C:\Users\predator\Desktop\new ims\Ims-desktop-application\supabase\variant_patch_chunks'
CHUNK_SIZE = 1500

os.makedirs(OUT_DIR, exist_ok=True)

wb = openpyxl.load_workbook(XLSX, read_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

barcode_idx = headers.index('Bar Code')
color_idx   = headers.index('Color')
size_idx    = headers.index('Size')

def sql_str(v):
    if v:
        return "'" + v.replace("'", "''") + "'"
    return 'NULL'

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    bc = str(row[barcode_idx]).strip() if row[barcode_idx] is not None else ''
    c  = str(row[color_idx]).strip().upper()  if row[color_idx]  is not None else ''
    s  = str(row[size_idx]).strip().upper()   if row[size_idx]   is not None else ''
    if not bc or (not c and not s):
        continue
    rows.append((bc, c, s))

total_chunks = math.ceil(len(rows) / CHUNK_SIZE)
print("Total rows:", len(rows), "  Chunks:", total_chunks)

for chunk_num in range(total_chunks):
    chunk = rows[chunk_num * CHUNK_SIZE : (chunk_num + 1) * CHUNK_SIZE]
    filename = os.path.join(OUT_DIR, f'patch_{chunk_num+1:02d}_of_{total_chunks:02d}.sql')
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f'-- Chunk {chunk_num+1}/{total_chunks}  |  rows {chunk_num*CHUNK_SIZE+1}-{chunk_num*CHUNK_SIZE+len(chunk)}\n')
        f.write('-- Run chunks in order (order does not affect correctness, all target barcode match)\n\n')
        f.write('UPDATE product_variants AS pv\n')
        f.write('SET\n')
        f.write('  color = COALESCE(v.color, pv.color),\n')
        f.write('  size  = COALESCE(v.size,  pv.size)\n')
        f.write('FROM (\n')
        f.write('  VALUES\n')
        for i, (bc, c, s) in enumerate(chunk):
            comma = '' if i == len(chunk) - 1 else ','
            f.write('    (' + sql_str(bc) + ', ' + sql_str(c) + ', ' + sql_str(s) + ')' + comma + '\n')
        f.write(') AS v(barcode, color, size)\n')
        f.write('WHERE pv.barcode = v.barcode;\n')
    size_kb = round(os.path.getsize(filename) / 1024, 1)
    print(f"  {os.path.basename(filename)}  ({len(chunk)} rows, {size_kb} KB)")

print("\nDone! Paste each file one at a time into the Supabase SQL editor.")
